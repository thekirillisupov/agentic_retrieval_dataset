"""Third-party ЕИС dumps (CSV/XLSX) -> the same document model as the XML export.

Getting a ЕИС token takes days and a Госуслуги-verified account, so this is the
path that works *today*: several third parties have published extracts of
zakupki.gov.ru that are still downloadable without credentials. They are
**record cards, not documents** — a row carries the notification's key fields
(объект закупки, заказчик, НМЦК, обеспечение, победитель), not its full text —
so a row yields three or four short sections where a real XML document yields a
dozen. Thinner, but real ЕИС prose with the template structure intact, and it
lands in exactly the format :mod:`arqg.zakupki.corpus` already chunks.

Profiles are for dumps whose columns were checked by hand; anything else goes
through ``generic`` with an explicit ``--column`` mapping. Unmapped columns are
ignored rather than guessed at — a mis-mapped ИНН would quietly poison the
entity mining downstream.

Licences differ per dump and are recorded in :data:`PROFILES` — check the one you
use before shipping a model trained on it.
"""
from __future__ import annotations

import csv
import os
import re
import sys
from dataclasses import dataclass, field
from typing import Any, Iterable, Iterator

from ..utils import log
from .parse import ProcurementDoc, Section

csv.field_size_limit(min(sys.maxsize, 2**31 - 1))

#: canonical field -> label used in the rendered section
FIELD_LABELS: dict[str, str] = {
    "doc_id": "Номер закупки",
    "doc_number": "Номер документа",
    "published": "Дата размещения",
    "law": "Закон",
    "procedure": "Способ определения поставщика",
    "url": "Ссылка на документ",
    "subject": "Наименование объекта закупки",
    "lot_name": "Наименование лота",
    "okpd2_code": "Код ОКПД2",
    "okpd2_name": "Наименование по ОКПД2",
    "item_description": "Описание позиций",
    "customer": "Заказчик",
    "customer_inn": "ИНН заказчика",
    "customer_region": "Регион заказчика",
    "region_code": "Код региона",
    "sponsor": "Организация, осуществляющая размещение",
    "price_start": "Начальная (максимальная) цена контракта",
    "price_final": "Цена контракта по результатам закупки",
    "security": "Обеспечение заявки",
    "advance": "Размер аванса",
    "currency": "Валюта",
    "small_business": "Закупка у субъектов малого предпринимательства",
    "phase": "Этап определения поставщика",
    "winner": "Победитель",
    "winner_inn": "ИНН победителя",
    "contract_date": "Дата заключения контракта",
    "contract_url": "Ссылка на контракт",
    "trade_date": "Дата проведения процедуры",
}

#: section title -> canonical fields, in the order they should be rendered
SECTION_SPEC: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("Общие сведения", ("doc_id", "doc_number", "published", "law", "procedure",
                        "trade_date", "url")),
    ("Объект закупки", ("subject", "lot_name", "okpd2_code", "okpd2_name",
                        "item_description")),
    ("Сведения о заказчике", ("customer", "customer_inn", "customer_region",
                              "region_code", "sponsor")),
    ("Цена и обеспечение", ("price_start", "currency", "security", "advance",
                            "small_business")),
    ("Результаты определения поставщика", ("phase", "winner", "winner_inn",
                                           "price_final", "contract_date",
                                           "contract_url")),
)

MONEY_FIELDS = {"price_start", "price_final", "security"}


@dataclass
class TableProfile:
    """How to read one published dump."""

    name: str
    source: str                       # where it came from, for the manifest
    licence: str                      # as declared by the publisher
    columns: dict[str, str] = field(default_factory=dict)   # canonical -> column
    delimiter: str = ","
    encoding: str = "utf-8"
    doc_type: str = "eisRecordCard"
    notes: str = ""
    #: Where to recover the procurement number when the id column is unusable —
    #: XLSX dumps routinely store a 19-digit number as a float and lose its tail.
    #: ``doc_id_from`` names another canonical field, ``doc_id_regex`` picks the
    #: number out of it.
    doc_id_from: str = ""
    doc_id_regex: str = r"(\d{19})"


PROFILES: dict[str, TableProfile] = {
    # kaggle datasets download dadalyndell/russian-biggest-government-procurement-contracts
    "kaggle_biggest": TableProfile(
        name="kaggle_biggest",
        source="kaggle.com/datasets/dadalyndell/russian-biggest-government-procurement-contracts",
        licence="ODC PDDL (public domain)",
        delimiter=",",
        notes="tenders over 500 mln RUB, 2014-2023, ~4.5k rows, winner included",
        columns={
            "doc_id": "tender_id", "subject": "tender_name",
            "price_start": "start_price", "security": "tender_security",
            "advance": "advance_money", "currency": "currency",
            "published": "publication_date", "phase": "selection_phase",
            "law": "legislation", "url": "url", "procedure": "procedure",
            "small_business": "for_small_business",
            "region_code": "customer_region_code", "customer_region": "customer_region",
            "customer": "customer_name", "customer_inn": "customer_inn",
            "winner": "winner_name", "winner_inn": "winner_inn",
            "price_final": "final_price",
        }),
    # kaggle datasets download mrmorj/zakupkihack-recsys  (train_data.csv)
    "zakupkihack": TableProfile(
        name="zakupkihack",
        source="kaggle.com/datasets/mrmorj/zakupkihack-recsys",
        licence="not declared by the publisher — check before redistribution",
        delimiter=";",
        notes="~0.5M+ lots, 2019-2020, 44/223-ФЗ, ОКПД2 + item descriptions, "
              "procurement numbers anonymised (pn_lot_*)",
        columns={
            "doc_id": "pn_lot_anon", "law": "fz", "region_code": "region_code",
            "published": "min_publish_date", "subject": "purchase_name",
            "lot_name": "lot_name", "price_start": "lot_price",
            "okpd2_code": "okpd2_code", "okpd2_name": "okpd2_names",
            "item_description": "item_descriptions",
        }),
    # huggingface.co/datasets/zavzyatiy/medicines_from_zakupki_gov_ru
    "hf_medicines": TableProfile(
        name="hf_medicines",
        source="huggingface.co/datasets/zavzyatiy/medicines_from_zakupki_gov_ru",
        licence="Apache-2.0",
        notes="drug procurement only (zakupki/tenders_farmcom_info.xlsx); the "
              "Num_trade column is a float and has lost digits — the real number "
              "is recovered from the trade URL",
        doc_id_from="url",
        columns={
            "doc_id": "Num_trade", "url": "Trade", "published": "Date_tpub",
            "trade_date": "Date_trade", "contract_url": "Contract",
            "contract_date": "Date_contract", "subject": "Name",
            "sponsor": "Sponsor", "customer": "Customer",
            "price_start": "IMCP", "price_final": "Price",
        }),
    "generic": TableProfile(
        name="generic", source="", licence="unknown",
        notes="supply --column canonical=source pairs yourself"),
}


# --------------------------------------------------------------------------- #
# reading
# --------------------------------------------------------------------------- #
def read_rows(path: str, profile: TableProfile) -> Iterator[dict[str, Any]]:
    """Yield rows from a ``.csv``/``.tsv`` or ``.xlsx`` file."""
    low = path.lower()
    if low.endswith((".xlsx", ".xlsm")):
        yield from _read_xlsx(path)
    else:
        with open(path, newline="", encoding=profile.encoding, errors="replace") as f:
            for row in csv.DictReader(f, delimiter=profile.delimiter):
                yield row


def _read_xlsx(path: str) -> Iterator[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as e:                            # pragma: no cover
        raise SystemExit("reading .xlsx needs openpyxl (pip install openpyxl)") from e
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(it)]
    for values in it:
        yield {h: v for h, v in zip(header, values)}
    wb.close()


def detect_profile(path: str) -> TableProfile | None:
    """Pick a profile by matching a file's header against the known column sets."""
    for profile in PROFILES.values():
        if not profile.columns:
            continue
        try:
            header = _header(path, profile)
        except (OSError, StopIteration):
            return None
        wanted = set(profile.columns.values())
        if wanted & header and len(wanted & header) >= max(3, len(wanted) // 2):
            log.info("zakupki: detected dump profile %r", profile.name)
            return profile
    return None


def _header(path: str, profile: TableProfile) -> set[str]:
    low = path.lower()
    if low.endswith((".xlsx", ".xlsm")):
        return set(next(iter(_read_xlsx(path))))
    with open(path, newline="", encoding=profile.encoding, errors="replace") as f:
        return set(next(csv.reader(f, delimiter=profile.delimiter)))


# --------------------------------------------------------------------------- #
# rendering
# --------------------------------------------------------------------------- #
_NUM_RE = re.compile(r"^-?\d+(?:[.,]\d+)?$")
_TS_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})(?:[ T](\d{2}):(\d{2}))?")
_LAWS = {"44fz": "44-ФЗ", "223fz": "223-ФЗ"}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).split())
    return "" if text.lower() in ("", "nan", "none", "null", "-") else text


def _format(field_name: str, value: Any) -> str:
    text = _clean(value)
    if not text:
        return ""
    if field_name == "law":
        return _LAWS.get(text.lower(), text)
    if text.upper() in ("TRUE", "FALSE"):
        return "да" if text.upper() == "TRUE" else "нет"
    m = _TS_RE.match(text)
    if m and field_name in ("published", "contract_date", "trade_date"):
        y, mo, d, hh, mm = m.groups()
        return f"{d}.{mo}.{y}" + (f" {hh}:{mm}" if hh else "")
    if field_name in MONEY_FIELDS and _NUM_RE.match(text.replace(" ", "")):
        num = float(text.replace(" ", "").replace(",", "."))
        whole, frac = f"{num:,.2f}".split(".")
        return f"{whole.replace(',', ' ')},{frac} руб."
    return text


def row_to_doc(row: dict[str, Any], profile: TableProfile,
               index: int = 0) -> ProcurementDoc | None:
    """One dump row -> a :class:`ProcurementDoc` with the usual sections."""
    values = {canon: _format(canon, row.get(col))
              for canon, col in profile.columns.items()}
    values = {k: v for k, v in values.items() if v}
    if not values:
        return None

    # Resolve the procurement number before rendering: it goes into the text as
    # well as the file name, and a dump whose id column is broken would collapse
    # thousands of distinct rows onto one document id.
    recovered = ""
    if profile.doc_id_from:
        m = re.search(profile.doc_id_regex, values.get(profile.doc_id_from, ""))
        recovered = m.group(1) if m else ""
    if recovered:
        values["doc_id"] = recovered
    doc_id = re.sub(r"[^\w.-]+", "_", values.get("doc_id", ""))[:80] or f"row{index:08d}"
    values.setdefault("doc_id", doc_id)

    sections: list[Section] = []
    for title, fields in SECTION_SPEC:
        lines = [f"{FIELD_LABELS.get(f, f)}: {values[f]}" for f in fields if f in values]
        if lines:
            sections.append(Section(title, lines))
    if not sections:
        return None

    subject = values.get("subject") or values.get("lot_name") or ""
    title = f"Карточка закупки № {doc_id}"
    if subject:
        title = f"{title}. {subject[:180]}"

    return ProcurementDoc(
        doc_type=profile.doc_type, doc_id=doc_id, title=title, sections=sections,
        file_ext=".txt",           # rendered from a table, not an EIS XML document
        meta={"source": profile.source, "licence": profile.licence,
              "profile": profile.name,
              "purchase_number": values.get("doc_id", ""),
              "published": values.get("published", ""),
              "customer": values.get("customer", ""),
              "subject": subject})


def iter_docs(path: str, profile: TableProfile, *,
              limit: int = 0) -> Iterator[ProcurementDoc]:
    """Parse a dump file into documents, skipping rows with nothing in them."""
    n_rows = n_docs = 0
    for i, row in enumerate(read_rows(path, profile)):
        n_rows += 1
        doc = row_to_doc(row, profile, i)
        if doc is None:
            continue
        n_docs += 1
        yield doc
        if limit and n_docs >= limit:
            break
    log.info("zakupki: %s -> %d documents from %d rows (%s)",
             os.path.basename(path), n_docs, n_rows, profile.name)


def parse_column_overrides(pairs: Iterable[str]) -> dict[str, str]:
    """``["subject=tender_name", "customer=org"]`` -> mapping, validated."""
    out: dict[str, str] = {}
    for pair in pairs:
        if "=" not in pair:
            raise SystemExit(f"--column expects canonical=source, got {pair!r}")
        canon, col = (p.strip() for p in pair.split("=", 1))
        if canon not in FIELD_LABELS:
            raise SystemExit(f"unknown field {canon!r}; known: {', '.join(sorted(FIELD_LABELS))}")
        out[canon] = col
    return out
